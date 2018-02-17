import pandas as pd
from matplotlib import pyplot as plt
import openpyxl

"""
clt_cog = pd.read_excel('clt cog.xlsx')
orl_cog = pd.read_excel('orl cog.xlsx')
ls_cog = pd.read_excel('ls cog.xlsx')
twd_cog = pd.read_excel('twd cog.xlsx')
hours_by_agent = pd.concat([clt_cog, orl_cog, ls_cog, twd_cog])
hours_by_agent.to_csv('hours_by_agent.csv', index=False)
print(hours_by_agent.head())
"""

hours_by_agent = pd.read_csv('hours_by_agent.csv')

"""
apt_1015 = pd.read_excel('October Appointments 10-15 to 10-20.xlsx')
apt_1022 = pd.read_excel('October Appointments 10-22 to 10-31.xlsx')
apt_111 = pd.read_excel('October Appointments 11-1 to 11-10.xlsx')
apt_1111 = pd.read_excel('October Appointments 11-11 to 11-20.xlsx')
apt_1121 = pd.read_excel('October Appointments 11-21 to 11-30.xlsx')
apt_121 = pd.read_excel('October Appointments 12-1 to 12-7.xlsx')
apt = pd.concat([apt_1015, apt_1022, apt_111, apt_1111, apt_1121, apt_121])
count_by_agent = apt.groupby('Agent_Name').Textbox20.count().reset_index()
count_by_agent.to_csv('count_by_agent.csv', index=False)
"""

"""
count_by_agent = pd.read_excel('count_by_agent.xlsx')
print(count_by_agent)
"""

"""
print(clt_cog.head())
print(ls_cog.head())
print(orl_cog.head())
print(twd_cog.head())
"""

urc = openpyxl.load_workbook('URC 10-15 to 12-7.xlsx')
hours_logged = urc.get_sheet_by_name('Sheet1')

def get_sec(time_str):
    """Returns time in seconds"""
    try:
        h, m, s = time_str.split(':')
        return int(h) * 3600 + int(m) * 60 + int(s)
    except ValueError:
        pass

dict1 = {}

for i in range(1, 500):
    if hours_logged.cell(row = i + 1, column = 4).value != None:
        e = str(hours_logged.cell(row = i + 9, column = 25).value)
        eout = get_sec(e)
        try:
            result = round(((aout + bout + cout + dout) / (eout)) * 100, 2)
        except TypeError:
            continue
        dict1[name] = result
